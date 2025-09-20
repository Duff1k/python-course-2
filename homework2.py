import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from datetime import datetime


def create_workbook(title):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Расписание'
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 18

    # Заголовки
    headers = ['Тренер', 'Вид спорта', 'Дата и время']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)

    wb.save(f'{title}.xlsx')


# Чтение и парсинг данных
data_about_workout = []
halls = set()

with open("trainings.txt", "r", encoding="utf-8") as data:
    for line in data:
        line = line.strip()
        if not line:
            continue

        parts = line.split(' | ')
        if len(parts) == 4:
            try:
                dt = datetime.strptime(parts[0], '%Y-%m-%d %H:%M')
                time_str = dt.strftime('%d.%m.%Y %H:%M')

                dict_with_data = {
                    'Дата и время': dt,  # сохраняем datetime для сортировки
                    'Время_строка': time_str,  # строка для отображения
                    'Спорт': parts[1],
                    'Тренер': parts[2].replace('Тренер: ', ''),
                    'Зал': parts[3]
                }
                data_about_workout.append(dict_with_data)
                halls.add(parts[3])

            except ValueError:
                print(f"Ошибка парсинга даты: {parts[0]}")

# Создаем файлы для каждого зала
for hall in halls:
    create_workbook(hall)

# Заполняем данные для каждого зала
for hall in halls:
    # Фильтруем данные для текущего зала
    hall_data = [item for item in data_about_workout if item['Зал'] == hall]

    # Сортируем по дате и времени
    hall_data_sorted = sorted(hall_data, key=lambda x: x['Дата и время'])

    # Открываем файл
    wb = load_workbook(f'{hall}.xlsx')
    ws = wb.active

    # Записываем данные
    for row, data_item in enumerate(hall_data_sorted, 2):  # начинаем с 2 строки
        ws[f'A{row}'] = data_item['Тренер']
        ws[f'B{row}'] = data_item['Спорт']
        ws[f'C{row}'] = data_item['Время_строка']

    # Сохраняем файл
    wb.save(f'{hall}.xlsx')

    print(f"Создан файл: {hall}.xlsx с {len(hall_data_sorted)} записями")