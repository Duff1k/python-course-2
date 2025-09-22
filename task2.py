import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

# сначала создам excel файлы с нужными настройками стиля
files = {}
for i in range(1, 5):
    wb = Workbook()
    ws = wb.active
    ws.title = "Расписание"

    headers = ["Тренер", "Вид спорта", "Дата и время"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    files[i] = {'wb': wb, 'ws': ws, 'data': []}

with open('trainings.txt', 'r', encoding='utf-8') as file:
    for line in file:
        if line.strip():
            line = line.split(' | ')
            i = int(line[3].split()[-1])
            files[i]['data'].append({
                'Дата и время': line[0],
                'Вид спорта': line[1],
                'Тренер': line[2].replace('Тренер: ', '')})

for i, file_data in files.items():
    file_data['data'].sort(key=lambda x: datetime.strptime(x['Дата и время'], '%Y-%m-%d %H:%M'))
    ws = file_data['ws']

    for row, item in enumerate(file_data['data'], 2):
        ws.cell(row=row, column=1, value=item['Тренер'])
        ws.cell(row=row, column=2, value=item['Вид спорта'])
        ws.cell(row=row, column=3, value=item['Дата и время'])

    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 16

    file_data['wb'].save(f'Зал {i}.xlsx')

# вообще удобнее сделать задание с помощью датафрейма для меня, поэтому я так тоже сделала
df = pd.read_csv('trainings.txt', sep=' \\| ', engine='python', header=None,
                 names=['Время', 'Вид спорта', 'Тренер', 'Зал'])
df['Время'] = pd.to_datetime(df['Время'], format='%Y-%m-%d %H:%M')
df['Тренер'] = df['Тренер'].str.replace('Тренер: ', '')
df['Зал'] = df['Зал'].str[-1].astype(int)
# print(df.info())

# сохраняю файлы excel
# только, чтобы выполнить условие про ширину колонок, придется открывать потом excel и менять стиль
for i in range(1, 5):
    df_gym = df[df['Зал'] == i].copy()
    df_gym = df_gym[['Тренер', 'Вид спорта', 'Время']]
    df_gym = df_gym.sort_values('Время')
    # df_gym.to_excel(f'Зал {i}.xlsx', index=False, sheet_name='Расписание')
