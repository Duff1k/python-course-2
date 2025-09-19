import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# Список для хранения данных
data = []

with open("trainings.txt", encoding="utf-8") as f:
    for i, line in enumerate(f, start=1):
        line = line.strip()
        if not line:
            continue

        # Разделяем по символу "|"
        parts = [p.strip() for p in line.split('|')]

        if len(parts) != 4:
            print(f"Строка {i} пропущена, ожидалось 4 поля: {line}")
            continue

        # Извлекаем данные
        date_time = parts[0]
        sport = parts[1]
        coach = parts[2].replace("Тренер:", "").strip()
        hall = parts[3]

        data.append({
            "Дата и время": date_time,
            "Вид спорта": sport,
            "Тренер": coach,
            "Зал": hall
        })

# Создаем DataFrame
df = pd.DataFrame(data)

# Получаем список залов
halls = df['Зал'].unique()

for hall in halls:
    df_hall = df[df['Зал'] == hall].copy()

    # Преобразуем дату и время
    df_hall['Дата и время'] = pd.to_datetime(df_hall['Дата и время'], errors='coerce')

    # Сортировка по дате и времени
    df_hall.sort_values(by='Дата и время', inplace=True)

    # Выбираем только нужные колонки
    df_hall = df_hall[['Тренер', 'Вид спорта', 'Дата и время']]

    # Имя файла Excel
    filename = f"{hall}.xlsx"

    # Сохраняем в Excel
    df_hall.to_excel(filename, index=False, sheet_name="Расписание")

    # Открываем файл для форматирования
    wb = load_workbook(filename)
    ws = wb["Расписание"]

    # Жирный шрифт для заголовков
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # Автоширина колонок
    for column_cells in ws.columns:
        max_length = 0
        column = column_cells[0].column
        for cell in column_cells:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(column)].width = max_length + 2

    wb.save(filename)

print("Файлы для каждого зала созданы и отформатированы.")
ы