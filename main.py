from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).resolve().parent
TXT_PATH = BASE_DIR / "trainings.txt"
OUT_DIR = BASE_DIR / 'out'
OUT_DIR.mkdir(parents=True, exist_ok=True)

TARGET_HALLS = [f'Зал {i}' for i in range(1,5)]
HEADERS = ["Тренер", "Вид спорта", "Дата и время"]
WIDTHS  = [22, 28, 20]

def parse_line(line: str):
    parts = [part.strip() for part in line.split('|')]
    if len(parts) != 4:
        return None
    dt_str, sport, coach_part, hall = parts
    coach = coach_part.replace('Тренер:', "").strip()
    try:
        dt = datetime.strptime(dt_str, '%Y-%m-%d %H:%M')
    except ValueError:
        dt = None
    return {"coach": coach, "sport": sport, "dt": dt, "dt_str": dt_str, "hall": hall}

by_hall = {h: [] for h in TARGET_HALLS}

with open(TXT_PATH, 'r', encoding='utf-8' ) as f:
    for row in f:
        line = row.strip()
        if not line:
            continue
        rec = parse_line(line)
        if not rec:
            continue
        if rec['hall'] in by_hall:
            by_hall[rec['hall']].append(rec)

for hall in TARGET_HALLS:
    by_hall[hall].sort(
        key=lambda r: (
            r['dt'] is None,
            r['dt'],
            r['coach']))

for hall in TARGET_HALLS:
    wb = Workbook()
    ws = wb.active
    ws.title = 'Расписание'

    ws.append(HEADERS)

    for col in range(1, len(HEADERS) + 1):
        ws.cell(row = 1, column=col).font = Font(bold = True)

    for rec in by_hall[hall]:
        dt_cell = rec["dt"] if rec["dt"] is not None else rec["dt_str"]
        ws.append([rec["coach"], rec["sport"], dt_cell])

    for i, w in enumerate(WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=3)
        if isinstance(cell.value, datetime):
            cell.number_format = "yyyy-mm-dd hh:mm"

    out_path = OUT_DIR / f"{hall}.xlsx"
    wb.save(out_path)