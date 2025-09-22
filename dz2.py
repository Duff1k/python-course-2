from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font

with open("trainings.txt", "r", encoding="utf-8") as file:
    lines = file.readlines()

rooms_data = {}

for line in lines:
    line = line.strip()
    if line:
        parts = line.split(' | ')
        if len(parts) >= 4:
            date_time_str = parts[0].strip()
            sport_type = parts[1].strip()
            coach = parts[2].replace('Тренер: ', '').strip()
            room = parts[3].strip()

            date_time = datetime.strptime(date_time_str, '%Y-%m-%d %H:%M')

            if room not in rooms_data:
                rooms_data[room] = []
            rooms_data[room].append({
                'Тренер': coach,
                'Вид спорта': sport_type,
                'Дата и время': date_time,
                'Дата для вывода': date_time.strftime('%d-%m-%Y %H:%M')
            })

for room_name, trainings in rooms_data.items():
    def get_date(training):
        return training['Дата и время']
    sorted_trainings = sorted(trainings, key=get_date)

    wb = Workbook()
    ws = wb.active
    ws.title = "Расписание"

    headers = ['Тренер', 'Вид спорта', 'Дата и время']

    ws.cell(row=1, column=1, value=headers[0])
    ws.cell(row=1, column=2, value=headers[1])
    ws.cell(row=1, column=3, value=headers[2])
    cell1 = ws.cell(row=1, column=1, value=headers[0])
    cell1.font = Font(bold=True)
    cell2 = ws.cell(row=1, column=1, value=headers[1])
    cell2.font = Font(bold=True)
    cell3 = ws.cell(row=1, column=1, value=headers[2])
    cell1.font = Font(bold=True)

    for row_num, training in enumerate(sorted_trainings, 2):
        ws.cell(row=row_num, column=1, value=training['Тренер'])
        ws.cell(row=row_num, column=2, value=training['Вид спорта'])
        ws.cell(row=row_num, column=3, value=training['Дата для вывода'])

    ws.column_dimensions['A'].width = 20  # Тренер
    ws.column_dimensions['B'].width = 25  # Вид спорта
    ws.column_dimensions['C'].width = 20  # Дата и время

    filename = f'{room_name}.xlsx'
    wb.save(filename)

print("Расписание по залам создано.")