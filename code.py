from openpyxl import Workbook
from openpyxl.styles import Font

halls = {}

with open('trainings.txt', 'r', encoding='utf-8') as f:
    for line in f:
        line = line.strip()
        if not line:
            continue

        parts = [part.strip() for part in line.split('|')]

        date_time, sport, coach, hall = parts
        coach = coach.replace("Тренер: ", "")

        if hall not in halls:
            halls[hall] = []

        halls[hall].append((coach, sport, date_time))

for hall, data in halls.items():
    data_sorted = sorted(data, key=lambda x: x[2])

    wb = Workbook()
    ws = wb.active
    ws.title = "Расписание"

    headers = ["Тренер", "Вид спорта", "Дата и время"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)

    for row_num, (coach, sport, date_time) in enumerate(data_sorted, 2):
        ws.cell(row=row_num, column=1, value=coach)
        ws.cell(row=row_num, column=2, value=sport)
        ws.cell(row=row_num, column=3, value=date_time)

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20

    wb.save(f"{hall}.xlsx")

