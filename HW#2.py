import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


# Файл построчно
with open('trainings.txt', "r", encoding="utf-8") as f:
    lines = [line.strip() for line in f if line.strip()]

# Парсинг строк
data = []
for line in lines:
    datetime_str, sport, coach, hall = [x.strip() for x in line.split("|")]
    coach = coach.replace("Тренер:", "").strip()
    data.append([coach, sport, datetime_str, hall])

df = pd.DataFrame(data, columns=["Тренер", "Вид спорта", "Дата и время", "Зал"])

df["Дата"] = pd.to_datetime(df["Дата и время"], format="%Y-%m-%d %H:%M")

# Сортировка
df = df.sort_values("Дата")

# Разбивка по залам
output_files = []
for hall in df["Зал"].unique():
    hall_df = df[df["Зал"] == hall].copy()
    hall_df = hall_df.drop(columns=["Зал", "Дата"])
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Расписание"
    
    headers = list(hall_df.columns)
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
    
    for row_num, row in enumerate(hall_df.values, 2):
        for col_num, value in enumerate(row, 1):
            ws.cell(row=row_num, column=col_num, value=value)

    for col_num, header in enumerate(headers, 1):
        max_length = max([len(str(ws.cell(row=row, column=col_num).value)) for row in range(1, ws.max_row+1)])
        ws.column_dimensions[get_column_letter(col_num)].width = max_length + 2
    
    hall_number = hall.replace("Зал ", "")
    wb.save(f'Зал {hall_number}.xlsx')

