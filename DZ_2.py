import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from datetime import datetime

with open('trainings.txt', encoding='utf-8') as f:
    lines = [line.strip() for line in f if line.strip()]

schedule = {1: [], 2: [], 3: [], 4: []}

for line in lines:
    dt_str, sport, trainer, hall = [x.strip() for x in line.split('|')]
    hall_num = int(hall.split()[-1])
    trainer_name = trainer.replace("Тренер: ", "")
    row = [trainer_name, sport, dt_str]
    schedule[hall_num].append((datetime.strptime(dt_str, "%Y-%m-%d %H:%M"), row))

for hall_num in range(1, 5):
    data = sorted(schedule[hall_num], key=lambda x: x[0])
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Расписание"
    headers = ["Тренер", "Вид спорта", "Дата и время"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
    for idx, (dt, row) in enumerate(data, start=2):
        for col, value in enumerate(row, start=1):
            ws.cell(row=idx, column=col, value=value)
    for col in range(1, 4):
        ws.column_dimensions[get_column_letter(col)].width = 22
    wb.save(f'Зал {hall_num}.xlsx')
