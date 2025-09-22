from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from datetime import datetime

with open("trainings.txt", "r", encoding="utf-8") as f:
    lines = f.readlines()

halls = {}

for line in lines:
    line = line.strip()
    if not line:
        continue
    parts = line.split(" | ")
    datetime_str, sport, coach, hall = parts
    coach = coach.replace("Тренер: ", "")   
    hall = hall.strip()                     
    if hall not in halls:
        halls[hall] = []
    halls[hall].append([coach, sport, datetime_str])

for hall, records in halls.items():
    records.sort(key=lambda x: datetime.strptime(x[2], "%Y-%m-%d %H:%M"))

    wb = Workbook()
    ws = wb.active
    ws.title = "Расписание"

    headers = ["Тренер", "Вид спорта", "Дата и время"]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in records:
        ws.append(row)

    for row in ws.iter_rows(min_row=2):  
        for cell in row:
            cell.alignment = Alignment(horizontal="left", vertical="center")

    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_len + 2

    wb.save(f"{hall}.xlsx")