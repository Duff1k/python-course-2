from openpyxl import Workbook
from openpyxl.styles import Font
import datetime
hall_data = {}
with open('trainings.txt', 'r', encoding='utf-8') as f:
    for line in f:
        line = line.strip()
        if not line:
            continue

        parts = [p.strip() for p in line.split('|')]
        date_time = parts[0]
        sport = parts[1]
        trainer = parts[2].replace('Тренер: ', '')
        hall_num = int(parts[3].split()[1])
        if hall_num not in hall_data:
            hall_data[hall_num] = []
        hall_data[hall_num].append((trainer, sport, date_time))
for hall_num in sorted(hall_data.keys()):
    wb = Workbook()
    ws = wb.active
    ws.title = "Расписание"
    headers = ["Тренер", "Вид спорта", "Дата и время"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
    data = hall_data[hall_num]
    data.sort(key=lambda x: datetime.datetime.strptime(x[2], '%Y-%m-%d %H:%M'))

    # Записываем данные
    for row_idx, (trainer, sport, dt) in enumerate(data, 2):
        ws.cell(row=row_idx, column=1, value=trainer)
        ws.cell(row=row_idx, column=2, value=sport)
        ws.cell(row=row_idx, column=3, value=dt)

    # Простая настройка ширины (можно убрать, если не критично)
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 20

    # Сохраняем
    wb.save(f"Зал {hall_num}.xlsx")
    print(f"✅ Файл 'Зал {hall_num}.xlsx' создан.")