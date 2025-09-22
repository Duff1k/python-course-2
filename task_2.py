from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
wb1 = Workbook()
wb2= Workbook()
wb3 = Workbook()
wb4 = Workbook()
ws1 = wb1.active
ws2 = wb2.active
ws3 = wb3.active
ws4 = wb4.active
ws1.title = "Расписание"
ws2.title = "Расписание"
ws3.title = "Расписание"
ws4.title = "Расписание"
headers = ["Тренер", "Вид спорта", "Дата и время"]
ws1.append(headers)
ws2.append(headers)
ws3.append(headers)
ws4.append(headers)

for cell in ws1[1]:
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal= "center", vertical = "center")
for cell in ws2[1]:
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal= "center", vertical = "center")
for cell in ws3[1]:
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal= "center", vertical = "center")
for cell in ws4[1]:
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal= "center", vertical = "center")
ws1.column_dimensions["A"].width=30
ws1.column_dimensions["B"].width=25
ws1.column_dimensions["C"].width=20
ws2.column_dimensions["A"].width=30
ws2.column_dimensions["B"].width=25
ws2.column_dimensions["C"].width=20
ws3.column_dimensions["A"].width=30
ws3.column_dimensions["B"].width=25
ws3.column_dimensions["C"].width=20
ws4.column_dimensions["A"].width=30
ws4.column_dimensions["B"].width=25
ws4.column_dimensions["C"].width=20
with open("trainings.txt", "r", encoding = "utf-8") as file:
    lines = file.readlines() 
    for i in range(len(lines)):
        line = lines[i].strip().split(" | ")
        if line[3] == "Зал 1":
            ws1.append([line[2], line[1], line[0]])
        elif line[3] == "Зал 2": 
            ws2.append([line[2], line[1], line[0]])
        elif line[3] == "Зал 3": 
            ws3.append([line[2], line[1], line[0]])
        else:
            ws4.append([line[2], line[1], line[0]]) 
wb1.save("Зал 1.xlsx")
wb2.save("Зал 2.xlsx")
wb3.save("Зал 3.xlsx")
wb4.save("Зал 4.xlsx")
#сортировка по дате и времени
for ws in [ws1, ws2, ws3, ws4]:
    schedule = list(ws.values)[1:]  
    sorted_schedule = sorted(schedule, key=lambda row: row[2]) 
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=3):
        for cell in row:
            cell.value = None 
    for row_data in sorted_schedule:
        ws.append(row_data) 
    