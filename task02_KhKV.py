import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from datetime import datetime

with open('trainings.txt', encoding='utf-8') as i:
    lines = [line.strip() for line in i if line.strip()]

schedule = {1: [], 2: [], 3: [], 4: []}

for line in lines:
    dt_tm, sport, trainer, gym = [x.strip() for x in line.split('|')]
    gym_num = int(gym.split()[-1])
    trainer_name = trainer.replace("Тренер: ", "")
    line = [trainer_name, sport, dt_tm]
    schedule[gym_num].append((datetime.strptime(dt_tm, "%Y-%m-%d %H:%M"), line))

for gym_num in range(1, 5):
    data = sorted(schedule[gym_num], key=lambda x: x[0])
    wb = openpyxl.Workbook()
    ws = wb.active
    headers = ["Тренер", "Вид спорта", "Дата и время"]
    for clmn, vl in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=clmn, value=vl)
        cell.font = Font(bold=True)
    for row_num, (dt, row) in enumerate(data, start=2):
        for clmn, vl in enumerate(row, start=1):
            ws.cell(row=row_num, column=clmn, value=vl)
    for clmn in range(1, 4):
        ws.column_dimensions[get_column_letter(clmn)].width = 20
    wb.save(f'Зал {gym_num}.xlsx')