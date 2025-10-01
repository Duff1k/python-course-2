# pip install pandas openpyxl

import pandas as pd
from datetime import datetime
import os
from openpyxl.styles import Border, Side

def parse_training_data(file_path):
    """Чтение и парсинг данных из файла"""
    trainings = []
    
    try:
        with open(file_path, 'r', encoding='utf-8') as file:
            for line in file:
                line = line.strip()
                if line:
                    parts = line.split(' | ')
                    if len(parts) == 4:
                        date_time_str, sport, trainer_info, hall = parts
                        
                        # Извлекаем имя тренера (убираем "Тренер: ")
                        trainer_name = trainer_info.replace('Тренер: ', '')
                        
                        # Преобразуем строку даты/времени в объект datetime
                        try:
                            date_time = datetime.strptime(date_time_str, '%Y-%m-%d %H:%M')
                        except ValueError:
                            continue
                        
                        trainings.append({
                            'Тренер': trainer_name,
                            'Вид спорта': sport,
                            'Дата и время': date_time,
                            'Зал': hall
                        })
    except FileNotFoundError:
        print(f"Ошибка: Файл не найден по пути {file_path}")
        print("Проверьте правильность пути к файлу")
        return None
    except Exception as e:
        print(f"Ошибка при чтении файла: {e}")
        return None
    
    return pd.DataFrame(trainings)

def add_borders_to_table(worksheet, num_rows, num_cols):
    """Добавляет черные границы ко всем ячейкам таблицы"""
    # Создаем стиль границы
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    # Применяем границы ко всем ячейкам таблицы
    for row in range(1, num_rows + 1):
        for col in range(1, num_cols + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.border = thin_border

def create_excel_files(df):
    """Создание Excel-файлов для каждого зала"""
    if df is None or df.empty:
        print("Нет данных для обработки")
        return
    
    halls = df['Зал'].unique()
    
    for hall in halls:
        # Фильтруем данные для текущего зала
        hall_data = df[df['Зал'] == hall].copy()
        
        # Сортируем по дате и времени
        hall_data = hall_data.sort_values('Дата и время')
        
        # Выбираем нужные колонки в правильном порядке
        hall_data = hall_data[['Тренер', 'Вид спорта', 'Дата и время']]
        
        # Создаем Excel файл
        file_name = f'{hall}.xlsx'
        
        try:
            with pd.ExcelWriter(file_name, engine='openpyxl') as writer:
                hall_data.to_excel(writer, sheet_name='Расписание', index=False)
                
                # Получаем рабочую книгу и лист для форматирования
                workbook = writer.book
                worksheet = writer.sheets['Расписание']
                
                # Делаем заголовки жирными
                for col_num, value in enumerate(hall_data.columns, 1):
                    cell = worksheet.cell(row=1, column=col_num)
                    cell.font = cell.font.copy(bold=True)
                
                # Устанавливаем ширину колонок
                column_widths = {
                    'Тренер': 20,
                    'Вид спорта': 25,
                    'Дата и время': 20
                }
                
                for col_num, column_name in enumerate(hall_data.columns, 1):
                    column_letter = chr(64 + col_num)
                    worksheet.column_dimensions[column_letter].width = column_widths[column_name]
                
                # Форматируем дату и время
                for row in range(2, len(hall_data) + 2):
                    cell = worksheet.cell(row=row, column=3)
                    cell.number_format = 'YYYY-MM-DD HH:MM'
                
                # Добавляем черные границы ко всей таблице
                num_rows = len(hall_data) + 1  # +1 для заголовка
                num_cols = len(hall_data.columns)
                add_borders_to_table(worksheet, num_rows, num_cols)
            
            print(f"Создан файл: {file_name}")
            
        except Exception as e:
            print(f"Ошибка при создании файла {file_name}: {e}")

def main():
    file_path = os.path.join(os.path.dirname(__file__), "trainings.txt")
    
    # Проверяем существование файла
    if not os.path.exists(file_path):
        print(f"Файл не найден по указанному пути: {file_path}")
        print("Пожалуйста, проверьте:")
        print("1. Правильность пути к файлу")
        print("2. Существование файла trainings.txt")
        print("3. Права доступа к файлу")
        return
    
    print(f"Чтение файла: {file_path}")
    
    # Чтение данных из файла
    df = parse_training_data(file_path)
    
    if df is not None:
        # Создание Excel файлов для каждого зала
        create_excel_files(df)
        
        print("\nГотово! Excel файлы успешно созданы для каждого зала")
        print("Созданные файлы находятся в текущей директории:")
        print(os.getcwd())
    else:
        print("Не удалось обработать данные из файла")

if __name__ == "__main__":
    main()
